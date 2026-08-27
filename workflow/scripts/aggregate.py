"""
Aggregate per-sample MLST call TSVs into:
    - summary  : one row per sample (wide format), joined with samplesheet metadata,
                 plus locus-balance columns and the basecalling model
    - long     : one row per (sample, locus) with allele/identity/coverage/flag/share
    - xlsx     : the same two views in one Excel workbook with QC-coded cell colours

The locus-balance and basecall-model helpers are pure functions so test/ can
import them; Snakemake state is read only at the bottom of the file.
"""
from pathlib import Path

import pandas as pd

# openpyxl is imported inside main(): the pure helpers below must stay importable
# by test/ without the Excel dependency.


def locus_shares(coverages: dict) -> dict:
    """Share (%) of a sample's on-target reads at each locus. Empty if no reads."""
    total = sum(float(v) for v in coverages.values())
    if total <= 0:
        return {l: 0.0 for l in coverages}
    return {l: 100.0 * float(v) / total for l, v in coverages.items()}


def balance_note(shares: dict, min_share_pct: float) -> str:
    """Name the loci whose read share is below min_share_pct, weakest first.

    A locus starved of reads caps the whole sample's QC no matter how deep the
    run is; the fix is the primer ratio at the bench, not a lower threshold.
    """
    low = sorted((s, l) for l, s in shares.items() if s < min_share_pct)
    if not low:
        return ""
    return "low_share:" + ",".join(f"{l}={s:.1f}%" for s, l in low)


def model_family(model: str) -> str:
    """Collapse a basecall/medaka model name to its chemistry family (r941, r104, r1041)."""
    m = (model or "").lower()
    for fam in ("r1041", "r10.4.1", "r104", "r10.4", "r941", "r9.4.1"):
        if fam in m:
            return fam.replace(".", "").replace("r1041", "r1041")
    return "unknown"


def model_mismatch(basecall_model: str, medaka_model: str) -> str:
    """Warn when reads were basecalled on a different chemistry than medaka assumes."""
    b, m = model_family(basecall_model), model_family(medaka_model)
    if b == "unknown" or m == "unknown" or b == m:
        return ""
    return f"basecall_model_mismatch:{basecall_model} polished with {medaka_model}"


def read_basecall_models(paths, loci_unused=None) -> dict:
    """{sample_id: model} from each results/<sample>/medaka/basecall_model.txt."""
    out = {}
    for p in paths:
        p = Path(p)
        sample = p.parent.parent.name
        try:
            out[sample] = p.read_text().strip() or "unknown"
        except OSError:
            out[sample] = "unknown"
    return out


def main():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    calls_in      = list(snakemake.input.calls)
    samplesheet   = snakemake.input.samplesheet
    summary_out   = snakemake.output.summary
    long_out      = snakemake.output.long
    xlsx_out      = snakemake.output.xlsx
    loci          = list(snakemake.params.loci)
    min_share_pct = float(snakemake.params.min_share_pct)
    medaka_model  = str(snakemake.params.medaka_model)
    basecall_map  = read_basecall_models(snakemake.input.basecall)

    QC_BG = {
        "PASS":         "E8F5E9",
        "NEW_ST":       "FFF8E1",
        "LOW_COVERAGE": "FFF1E0",
        "NEW_ALLELE":   "FCE4EC",
        "FAIL":         "FFEBEE",
    }
    FLAG_BG = {
        "known":      "E8F5E9",
        "new_allele": "FCE4EC",
        "no_hit":     "FFEBEE",
    }

    # ---- wide summary ----
    # Force sample_id (and all allele columns) to string. Otherwise pandas infers
    # numeric dtype for sheets whose sample_ids happen to be all digits, and the
    # samplesheet-vs-calls merge below silently fails with blank qc_label.
    _str_cols = {"sample_id": str}
    for _l in loci:
        _str_cols[_l] = str

    frames = [pd.read_csv(p, sep="\t", dtype=_str_cols) for p in calls_in]
    wide = pd.concat(frames, ignore_index=True)

    # Join samplesheet metadata (keep original samplesheet columns left of wide)
    meta = pd.read_csv(samplesheet, dtype={"sample_id": str})
    wide = meta.merge(wide, on="sample_id", how="left")

    # ---- locus balance + basecalling model (per sample) ----
    _shares = {}
    for _, r in wide.iterrows():
        sid = r["sample_id"]
        cov = {l: pd.to_numeric(r.get(f"{l}_coverage"), errors="coerce") or 0 for l in loci}
        sh = locus_shares(cov)
        _shares[sid] = sh
        for l in loci:
            wide.loc[wide["sample_id"] == sid, f"{l}_share_pct"] = round(sh[l], 2)
        wide.loc[wide["sample_id"] == sid, "locus_balance_note"] = balance_note(sh, min_share_pct)
        bc = basecall_map.get(sid, "unknown")
        wide.loc[wide["sample_id"] == sid, "basecall_model"] = bc
        wide.loc[wide["sample_id"] == sid, "model_note"] = model_mismatch(bc, medaka_model)

    Path(summary_out).parent.mkdir(parents=True, exist_ok=True)
    wide.to_csv(summary_out, sep="\t", index=False)

    # ---- long format: one row per (sample, locus) ----
    long_rows = []
    for _, r in wide.iterrows():
        for l in loci:
            long_rows.append({
                "sample_id": r["sample_id"],
                "run_id":    r.get("run_id", ""),
                "locus":     l,
                "allele":    r.get(l, "-"),
                "identity":  r.get(f"{l}_identity", 0),
                "coverage":  r.get(f"{l}_coverage", 0),
                "flag":      r.get(f"{l}_flag", "no_hit"),
                "share_pct": _shares.get(r["sample_id"], {}).get(l, 0.0),
            })
    long_df = pd.DataFrame(long_rows)
    long_df.to_csv(long_out, sep="\t", index=False)


    # ----------------------------- XLSX -----------------------------
    def autofit(ws, df: pd.DataFrame):
        for idx, col in enumerate(df.columns, 1):
            width = max(12, min(32, len(str(col)) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width


    def write_sheet(ws, df: pd.DataFrame, row_fill_by_col: str | None,
                    flag_cols: dict | None):
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center")
        for idx, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=idx, value=col)
            c.font = header_font
            c.alignment = header_align

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            row_fill = None
            if row_fill_by_col:
                qc = row[row_fill_by_col]
                if qc in QC_BG:
                    row_fill = PatternFill("solid", fgColor=QC_BG[qc])
            for c_idx, col in enumerate(df.columns, 1):
                val = row[col]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if row_fill is not None:
                    cell.fill = row_fill
            if flag_cols:
                for locus, col_idx in flag_cols.items():
                    flag_val = row.get(f"{locus}_flag", "")
                    if flag_val in FLAG_BG:
                        ws.cell(row=r_idx, column=col_idx).fill = PatternFill(
                            "solid", fgColor=FLAG_BG[flag_val]
                        )
        ws.freeze_panes = "A2"
        autofit(ws, df)


    wb = Workbook()

    # Sheet 1: Summary (QC-coloured rows + per-locus flag cell highlighting)
    ws1 = wb.active
    ws1.title = "Summary"
    flag_cols = {l: wide.columns.get_loc(f"{l}_flag") + 1
                 for l in loci if f"{l}_flag" in wide.columns}
    write_sheet(ws1, wide, row_fill_by_col="qc_label", flag_cols=flag_cols)

    # Sheet 2: Per-locus long format (flag cell coloured per row)
    ws2 = wb.create_sheet("Per-locus")
    long_flag_col = long_df.columns.get_loc("flag") + 1 if "flag" in long_df.columns else None
    header_font = Font(bold=True)
    for idx, col in enumerate(long_df.columns, 1):
        c = ws2.cell(row=1, column=idx, value=col)
        c.font = header_font
    for r_idx, (_, row) in enumerate(long_df.iterrows(), start=2):
        for c_idx, col in enumerate(long_df.columns, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            ws2.cell(row=r_idx, column=c_idx, value=val)
        if long_flag_col:
            flag_val = row.get("flag", "")
            if flag_val in FLAG_BG:
                ws2.cell(row=r_idx, column=long_flag_col).fill = PatternFill(
                    "solid", fgColor=FLAG_BG[flag_val]
                )
    ws2.freeze_panes = "A2"
    autofit(ws2, long_df)

    wb.save(xlsx_out)


if "snakemake" in globals():  # pragma: no cover - Snakemake entry point
    main()
