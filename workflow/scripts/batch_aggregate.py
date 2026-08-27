#!/usr/bin/env python3
"""
Cross-run aggregator for nanotyper.

Concatenates every per-run mlst_summary.tsv into one wide table, then
produces:
    combined_summary.tsv    one row per sample across all runs, with run_folder column
    combined_summary.xlsx   (if openpyxl is installed) same data with QC-coloured rows
    st_distribution.tsv     ST × run_folder pivot (how often each ST appears per run)
    qc_by_run.tsv           QC label × run_folder pivot
    replicates.tsv          biological_id × ST groups that disagree (if biological_id used)
"""
import argparse
import sys
from pathlib import Path

import pandas as pd

QC_BG = {
    "PASS":         "E8F5E9",
    "NEW_ST":       "FFF8E1",
    "LOW_COVERAGE": "FFF1E0",
    "NEW_ALLELE":   "FCE4EC",
    "FAIL":         "FFEBEE",
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis-root", type=Path, required=True,
                    help="(informational) root of per-run analyses")
    ap.add_argument("--output", type=Path, required=True,
                    help="directory to write combined outputs into")
    ap.add_argument("summaries", nargs="+",
                    help="paths to per-run mlst_summary.tsv files")
    args = ap.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)

    # --- load and tag each per-run summary ---
    frames = []
    for p in args.summaries:
        p = Path(p)
        # analysis folder name = 2 levels up (…/<run>/results/mlst_summary.tsv)
        run_folder = p.parent.parent.name
        df = pd.read_csv(p, sep="\t", dtype=str)
        df.insert(0, "run_folder", run_folder)
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True)

    # --- 1) combined wide summary ---
    tsv_path = args.output / "combined_summary.tsv"
    combined.to_csv(tsv_path, sep="\t", index=False)
    print(f"  wrote {tsv_path}  ({len(combined):,} rows, {len(args.summaries)} runs)")

    # --- 2) xlsx with QC-coloured rows ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Combined"

        for col_idx, col in enumerate(combined.columns, 1):
            c = ws.cell(row=1, column=col_idx, value=col)
            c.font = Font(bold=True)

        qc_col = combined.columns.get_loc("qc_label") + 1 if "qc_label" in combined.columns else None
        for r_idx, (_, row) in enumerate(combined.iterrows(), 2):
            fill = None
            if qc_col is not None:
                qc = row.get("qc_label", "")
                if qc in QC_BG:
                    fill = PatternFill("solid", fgColor=QC_BG[qc])
            for c_idx, col in enumerate(combined.columns, 1):
                val = row[col]
                cell = ws.cell(row=r_idx, column=c_idx, value="" if pd.isna(val) else val)
                if fill is not None:
                    cell.fill = fill

        for col_idx, col in enumerate(combined.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                max(12, min(30, len(str(col)) + 2))
        ws.freeze_panes = "A2"

        xlsx_path = args.output / "combined_summary.xlsx"
        wb.save(xlsx_path)
        print(f"  wrote {xlsx_path}")
    except ImportError:
        print("  (openpyxl not installed — skipped .xlsx output)")

    # --- 3) ST × run pivot ---
    if "ST" in combined.columns:
        called = combined[combined["ST"].astype(str).ne("-") & combined["ST"].notna()].copy()
        if not called.empty:
            pivot = (called.groupby(["ST", "run_folder"]).size()
                           .unstack(fill_value=0)
                           .sort_index())
            pivot["_total"] = pivot.sum(axis=1)
            pivot = pivot.sort_values("_total", ascending=False)
            pivot_path = args.output / "st_distribution.tsv"
            pivot.to_csv(pivot_path, sep="\t")
            print(f"  wrote {pivot_path}  ({len(pivot)} unique STs)")

    # --- 4) QC summary per run ---
    if "qc_label" in combined.columns:
        qc = (combined.groupby(["run_folder", "qc_label"]).size()
                      .unstack(fill_value=0)
                      .reindex(columns=["PASS","NEW_ST","LOW_COVERAGE","NEW_ALLELE","FAIL"],
                               fill_value=0))
        qc["total"] = qc.sum(axis=1)
        qc_path = args.output / "qc_by_run.tsv"
        qc.to_csv(qc_path, sep="\t")
        print(f"  wrote {qc_path}")
        print("\n  QC label counts by run:")
        print(qc.to_string().replace("\n", "\n  "))

    # --- 5) replicate disagreements (if biological_id used) ---
    disagree_df = None
    if "biological_id" in combined.columns:
        has_bio = combined[combined["biological_id"].notna() &
                           combined["biological_id"].astype(str).str.strip().ne("")]
        if not has_bio.empty:
            groups = (has_bio.groupby("biological_id")["ST"]
                              .agg(lambda s: sorted(set(str(x) for x in s))))
            # Replicates with >1 unique ST (ignoring "-" = not called)
            disagree = groups[groups.apply(
                lambda lst: len([x for x in lst if x not in ("-", "nan")]) > 1
            )]
            rep_path = args.output / "replicates.tsv"
            disagree.apply(lambda lst: ",".join(lst)).rename("ST_values") \
                    .to_csv(rep_path, sep="\t")
            print(f"  wrote {rep_path}  ({len(disagree)} replicate group(s) with disagreement)")
            if len(disagree) > 0:
                disagree_df = disagree.apply(lambda lst: ",".join(lst)).reset_index()
                disagree_df.columns = ["biological_id", "ST_values"]

    # --- 6) self-contained interactive HTML report ---
    try:
        html_path = args.output / "combined_report.html"
        write_html_report(
            output=html_path,
            combined=combined,
            qc_by_run=qc if "qc_label" in combined.columns else None,
            st_pivot=pivot if "ST" in combined.columns and not called.empty else None,
            disagree_df=disagree_df,
        )
        print(f"  wrote {html_path}")
    except ImportError as e:
        print(f"  (skipped HTML report: {e}; install plotly with `mamba install plotly`)")


# ======================================================================
# HTML report
# ======================================================================
QC_COLOURS = {
    "PASS":         "#28a745",
    "NEW_ST":       "#ffc107",
    "LOW_COVERAGE": "#fd7e14",
    "NEW_ALLELE":   "#e83e8c",
    "FAIL":         "#dc3545",
}


def _qc_level_colour(rate: float) -> tuple:
    """Banner gradient for the overall pass rate."""
    if rate >= 80:
        return "#28a745", "#1e7e34", "✓"
    if rate >= 50:
        return "#fd7e14", "#d06400", "⚠"
    return "#dc3545", "#a71d2a", "✗"


def write_html_report(output, combined, qc_by_run, st_pivot, disagree_df):
    import plotly.graph_objects as go
    from plotly.offline import plot as plot_to_div

    total_samples = len(combined)
    pass_samples  = int((combined.get("qc_label") == "PASS").sum())
    need_attn     = int(combined.get("qc_label").isin(["FAIL", "NEW_ALLELE"]).sum()) \
                      if "qc_label" in combined.columns else 0
    pass_rate     = 100 * pass_samples / max(total_samples, 1)
    n_runs        = combined["run_folder"].nunique()
    unique_sts    = 0
    if "ST" in combined.columns:
        known = combined[combined["qc_label"].isin(["PASS","LOW_COVERAGE"])] \
                        if "qc_label" in combined.columns else combined
        unique_sts = known["ST"].dropna().astype(str).replace("-", pd.NA).dropna().nunique()

    g1, g2, banner_icon = _qc_level_colour(pass_rate)
    if pass_rate >= 80:
        verdict = f"Overall run quality is strong — {pass_samples} of {total_samples} samples ({pass_rate:.0f}%) fully typed."
    elif pass_rate >= 50:
        verdict = f"Mixed quality — {pass_samples} of {total_samples} ({pass_rate:.0f}%) fully typed; review {total_samples - pass_samples} samples below."
    else:
        verdict = f"Quality needs attention — only {pass_samples} of {total_samples} ({pass_rate:.0f}%) fully typed."

    charts = []

    # Per-run QC stacked bar (by count)
    if qc_by_run is not None and not qc_by_run.empty:
        qcr = qc_by_run.drop(columns=["total"], errors="ignore")
        fig = go.Figure()
        for label in ["PASS","LOW_COVERAGE","NEW_ST","NEW_ALLELE","FAIL"]:
            if label in qcr.columns:
                fig.add_trace(go.Bar(
                    name=label, x=qcr.index, y=qcr[label],
                    marker_color=QC_COLOURS.get(label, "#666"),
                ))
        fig.update_layout(
            barmode="stack",
            title="QC outcomes per run (count)",
            xaxis_title="Run", yaxis_title="Samples",
            template="simple_white", height=380,
        )
        charts.append(("QC by run", plot_to_div(fig, include_plotlyjs=False, output_type="div")))

    # Top STs across all runs
    if "ST" in combined.columns:
        tops = combined[combined["ST"].astype(str).replace("-", pd.NA).notna()] \
                   .groupby("ST").size().reset_index(name="n") \
                   .sort_values("n", ascending=True).tail(20)
        if not tops.empty:
            fig = go.Figure(go.Bar(
                x=tops["n"], y="ST-" + tops["ST"].astype(str),
                orientation="h",
                marker=dict(color=tops["n"], colorscale="Viridis"),
                text=tops["n"], textposition="outside",
            ))
            fig.update_layout(
                title="Top sequence types (across all runs)",
                xaxis_title="Samples", yaxis_title="",
                template="simple_white", height=max(360, 18 * len(tops) + 60),
            )
            charts.append(("Top STs", plot_to_div(fig, include_plotlyjs=False, output_type="div")))

    # ST × run heatmap (if enough data)
    if st_pivot is not None:
        mat = st_pivot.drop(columns=["_total"], errors="ignore")
        if len(mat) > 0 and mat.shape[1] > 1:
            # Show top 30 STs by total
            show = mat.head(30)
            fig = go.Figure(go.Heatmap(
                z=show.values, x=list(show.columns), y=["ST-" + str(s) for s in show.index],
                colorscale="Viridis", colorbar=dict(title="Count"),
                hovertemplate="ST=%{y}<br>Run=%{x}<br>Count=%{z}<extra></extra>",
            ))
            fig.update_layout(
                title="ST distribution across runs (top 30 STs)",
                template="simple_white",
                height=max(420, 18 * len(show) + 80),
                xaxis=dict(tickangle=-30),
            )
            charts.append(("ST × run", plot_to_div(fig, include_plotlyjs=False, output_type="div")))

    # Combined sample table (HTML, with DataTables.js from CDN)
    table_cols = [c for c in
                  ["run_folder","sample_id","biological_id","run_id","barcode",
                   "sample_type","ST","qc_label","qc_notes"]
                  if c in combined.columns]
    table_html = combined[table_cols].to_html(
        index=False, classes="display compact cell-border", table_id="samples", border=0,
        escape=False,
    )

    # Replicate disagreements table (if any)
    disagree_html = ""
    if disagree_df is not None and len(disagree_df) > 0:
        disagree_html = (
            '<h2>Replicates that disagree</h2>'
            '<p>Samples sharing a <code>biological_id</code> whose calls don\'t all match. Review these.</p>'
            + disagree_df.to_html(index=False, classes="display compact cell-border",
                                  table_id="replicates", border=0, escape=False)
        )

    charts_html = "\n".join(
        f'<section><h2>{title}</h2>{html}</section>' for title, html in charts
    )

    # Build the page
    now_str = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    html_doc = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>nanotyper — cross-run report</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         margin: 0; padding: 0 24px 60px; color: #222; max-width: 1200px; margin: 0 auto; }}
  h1 {{ margin-top: 32px; }}
  h2 {{ margin-top: 38px; border-bottom: 1px solid #e2e8f0; padding-bottom: 4px; }}
  .banner {{ color: white; padding: 24px 28px; border-radius: 10px;
             margin: 18px 0 24px; box-shadow: 0 4px 10px rgba(0,0,0,0.12);
             background: linear-gradient(135deg, {g1} 0%, {g2} 100%); }}
  .banner .verdict {{ font-size: 1.55em; font-weight: 700; margin-bottom: 4px; }}
  .banner .sub     {{ opacity: 0.92; font-size: 1em; }}
  .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                 gap: 14px; margin: 12px 0 24px; }}
  .metric {{ color: white; padding: 16px 20px; border-radius: 10px;
             box-shadow: 0 3px 6px rgba(0,0,0,0.08); }}
  .metric .label {{ opacity: 0.9; font-size: 1em; }}
  .metric .value {{ font-size: 2.4em; font-weight: 700; margin: 4px 0; line-height: 1; }}
  .metric .sub   {{ opacity: 0.85; font-size: 0.9em; }}
  footer {{ text-align: center; margin-top: 40px; color: #64748b; font-size: 0.9em;
            border-top: 1px solid #e2e8f0; padding-top: 14px; }}
  table.display {{ width: 100% !important; font-size: 0.92em; }}
</style>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<script src="https://code.jquery.com/jquery-3.7.1.slim.min.js"></script>
<script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
</head>
<body>

<h1>nanotyper — cross-run report</h1>

<div class="banner">
  <div class="verdict">{banner_icon} &nbsp; {verdict}</div>
  <div class="sub">{n_runs} run(s) · {total_samples:,} samples · generated {now_str}</div>
</div>

<div class="metric-grid">
  <div class="metric" style="background:linear-gradient(135deg,{g1},{g2});">
    <div class="label">Fully typed (PASS)</div>
    <div class="value">{pass_samples}</div>
    <div class="sub">{pass_rate:.1f}% of total</div>
  </div>
  <div class="metric" style="background:linear-gradient(135deg,#6b7280,#4b5563);">
    <div class="label">Total samples</div>
    <div class="value">{total_samples}</div>
    <div class="sub">across {n_runs} runs</div>
  </div>
  <div class="metric" style="background:linear-gradient(135deg,#5a3d8c,#3d2463);">
    <div class="label">Unique STs</div>
    <div class="value">{unique_sts}</div>
    <div class="sub">PASS + LOW_COVERAGE</div>
  </div>
  <div class="metric" style="background:linear-gradient(135deg,{'#dc3545' if need_attn else '#6b7280'},{'#a71d2a' if need_attn else '#4b5563'});">
    <div class="label">Need attention</div>
    <div class="value">{need_attn}</div>
    <div class="sub">FAIL + NEW_ALLELE</div>
  </div>
</div>

{charts_html}

{disagree_html}

<h2>All samples (searchable)</h2>
{table_html}

<script>
$(function() {{
  $('#samples').DataTable({{ pageLength: 25, scrollX: true }});
  $('#replicates').DataTable({{ pageLength: 10, searching: false, paging: false, info: false }});
}});
</script>

<footer>
  nanotyper · cross-run aggregate · {now_str}
</footer>
</body>
</html>
"""
    output.write_text(html_doc)


if __name__ == "__main__":
    main()
